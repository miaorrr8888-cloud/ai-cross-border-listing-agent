"""商品输入采集：JSON / CSV / Excel 读取 + 批量生成。

读取列（兼容中英文表头）：
  sku, title(标题), images(图片, 逗号或换行分隔), price(价格=cost_price),
  currency(货币), url(来源链接), platform(平台)

批量生成：每条商品走一遍 dry-run 流水线，写入 output/products/<sku>.json。
"""
from __future__ import annotations

import csv
import json
import os
from typing import List

from src.models.product import Product
from src.utils.common import save_json


def from_json(path: str) -> List[Product]:
    """从 JSON 文件读取一个或多个商品，返回 Product 列表。"""
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    items = data if isinstance(data, list) else [data]
    return [Product.from_dict(i) for i in items]


def from_csv(path: str) -> List[Product]:
    """从 CSV 读取商品（标准库 csv，无额外依赖）。"""
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    return [_row_to_product(r) for r in rows]


def from_excel(path: str) -> List[Product]:
    """从 Excel(.xlsx/.xls) 读取商品（需要 openpyxl）。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError("读取 Excel 需要安装 openpyxl：pip install openpyxl")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    products = []
    for r in rows[1:]:
        d = {header[i]: (r[i] if i < len(r) else None) for i in range(len(header))}
        products.append(_row_to_product(d))
    return products


def read_products(path: str) -> List[Product]:
    """按扩展名分派到对应读取器。"""
    if path.endswith(".json"):
        return from_json(path)
    if path.endswith(".csv"):
        return from_csv(path)
    if path.endswith((".xlsx", ".xls")):
        return from_excel(path)
    raise ValueError(f"不支持的文件格式：{path}（支持 .json / .csv / .xlsx）")


def _split_images(value) -> List[str]:
    if not value:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    return [v.strip() for v in str(value).replace("\n", ",").split(",") if v.strip()]


def _to_float(value) -> float:
    try:
        return float(str(value).strip()) if value not in (None, "") else 0.0
    except ValueError:
        return 0.0


def _row_to_product(d: dict) -> Product:
    d = {str(k).lower() if k else "": v for k, v in d.items()}
    return Product(
        source_url=str(d.get("url") or d.get("source_url") or ""),
        source_platform=str(d.get("platform") or d.get("source_platform") or "generic"),
        title_cn=str(d.get("title") or d.get("title_cn") or ""),
        title_en=str(d.get("title_en") or ""),
        images=_split_images(d.get("images") or d.get("image")),
        sku=str(d.get("sku") or ""),
        brand=str(d.get("brand") or ""),
        cost_price=_to_float(d.get("price") or d.get("cost_price")),
        currency=str(d.get("currency") or "RMB"),
        keywords=_split_images(d.get("keywords")),
    )


def _safe(name: str) -> str:
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in name)


def batch_generate(
    products: List[Product],
    cfg: dict,
    out_dir: str = "output/products",
) -> List[dict]:
    """对每条商品跑 dry-run 流水线，逐条写入 <sku>.json，返回汇总列表。"""
    from src.pipeline import run

    # 相对路径统一解析到项目根目录，避免受运行目录影响
    if not os.path.isabs(out_dir):
        root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        out_dir = os.path.join(root, out_dir)
    os.makedirs(out_dir, exist_ok=True)
    summary = []
    for i, p in enumerate(products):
        out = run(p, cfg)
        sku = p.sku or f"item_{i + 1}"
        path = os.path.join(out_dir, f"{_safe(sku)}.json")
        save_json(out, path)
        summary.append({
            "sku": sku,
            "path": path,
            "score": out["listing_check"]["score"],
        })
    return summary
